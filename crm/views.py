from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps
from io import StringIO

from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.management import call_command
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .forms import (
    ApparatusForm,
    ChildForm,
    CompetitionEntryForm,
    CompetitionForm,
    ExpenseForm,
    LeadForm,
    ManagerTaskForm,
    NewcomerForm,
    ProfileForm,
    ReminderForm,
    RevenueTargetForm,
    StaffCreateForm,
    StyledPasswordChangeForm,
    SubscriptionForm,
    TariffForm,
)
from .models import (
    Apparatus,
    ApparatusScore,
    Attendance,
    AuditEvent,
    Child,
    Competition,
    CompetitionEntry,
    Expense,
    Group,
    Lead,
    ManagerTask,
    Newcomer,
    Payment,
    Reminder,
    RevenueTarget,
    Role,
    SalaryPayout,
    StaffProfile,
    Subscription,
    Tariff,
    Trainer,
    user_role,
)


PAGE_META = {
    "attendance": ("Табель", "Отмечайте посещения прямо в таблице"),
    "statistics": ("Статистика", "Главные показатели клуба"),
    "payments": ("Продления", "Кому пора напомнить об оплате"),
    "expenses": ("Расходы", "Бытовые закупки и другие расходы"),
    "competitions": ("Соревнования", "Баллы, места и история выступлений"),
    "notifications": ("Уведомления", "Задачи и события, требующие внимания"),
    "boss": ("Для руководителя", "Выручка, KPI и работа команды"),
    "profile": ("Мой профиль", "Личные данные и безопасность"),
    "applications": ("Заявки", "Все обращения из рекламы, сайта и звонков"),
    "newcomers": ("Новички", "Пробные занятия и переход к оплате"),
    "calendar": ("Календарь", "Напоминания и задачи команды"),
    "search": ("Поиск", "Спортсмены, заявки, группы и контакты"),
}


def role_required(min_rank):
    def decorator(view):
        @wraps(view)
        @login_required
        def wrapped(request, *args, **kwargs):
            if {Role.MANAGER: 0, Role.SENIOR: 1, Role.BOSS: 2}[user_role(request.user)] < min_rank:
                return HttpResponseForbidden("Недостаточно прав")
            return view(request, *args, **kwargs)

        return wrapped

    return decorator


def page_context(request, page, **extra):
    title, subtitle = PAGE_META[page]
    context = {
        "page": page,
        "title": title,
        "subtitle": subtitle,
        "current_role": user_role(request.user),
        "is_boss": user_role(request.user) == Role.BOSS,
        "is_senior": user_role(request.user) in (Role.SENIOR, Role.BOSS),
    }
    context.update(extra)
    return context


def log_action(request, action, obj, description):
    AuditEvent.objects.create(
        actor=request.user,
        action=action,
        object_type=obj.__class__.__name__ if obj else "",
        object_id=str(obj.pk) if obj and obj.pk else "",
        description=description,
    )


def login_page(request):
    if request.user.is_authenticated:
        return redirect("attendance")
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user and user.is_active:
            login(request, user)
            request.session.set_expiry(1209600 if request.POST.get("remember_me") else 0)
            return redirect(request.GET.get("next") or "attendance")
        messages.error(request, "Неверный логин или пароль")
    return render(request, "crm/login.html")


from datetime import timedelta, datetime
import calendar
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Group, ScheduleSlot, Child

from datetime import timedelta, datetime
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Group, ScheduleSlot, Child


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import timedelta, datetime
from .models import *
from .forms import ChildForm

def generate_class_dates(group, start_date, limit=60):
    """Генерирует список дат занятий"""
    slots = ScheduleSlot.objects.filter(group=group).order_by('weekday', 'start_time')
    if not slots:
        return []
    
    dates = []
    current = start_date
    end_limit = start_date + timedelta(days=180)
    
    while current <= end_limit and len(dates) < limit:
        for slot in slots:
            if current.weekday() == slot.weekday:
                dates.append(current)
        current += timedelta(days=1)
    
    return sorted(list(set(dates)))


@login_required
def logout_page(request):
    logout(request)
    return redirect("login")

def attendance_view(request):
    group_id = request.GET.get('group_id')
    ref_date_str = request.GET.get('ref_date')
    sort_by = request.GET.get('sort', 'name')
    show_archived = request.GET.get('show_archived') == '1'

    # 1. Группа
    if group_id:
        group = get_object_or_404(Group, id=group_id)
    else:
        group = Group.objects.filter(is_active=True).first()
        if not group:
            return render(request, 'crm/attendance.html', {
                'groups': Group.objects.none(),
                'selected_group': None,
            })

    trainer = group.trainer if group else None
    today = timezone.localdate()

    # 2. Опорная дата
    if ref_date_str:
        try:
            ref_date = datetime.strptime(ref_date_str, '%Y-%m-%d').date()
        except ValueError:
            ref_date = today
    else:
        ref_date = today

    # 3. Генерируем даты
    start_of_week = ref_date - timedelta(days=ref_date.weekday())
    all_class_dates = generate_class_dates(group, start_of_week, limit=60)

    if not all_class_dates:
        return render(request, 'crm/attendance.html', {
            'groups': Group.objects.filter(is_active=True),
            'selected_group': group,
            'week_data': [],
            'children_data': [],
            'ref_date': ref_date,
            'error': 'Нет расписания'
        })

    # 4. Находим индекс
    current_index = 0
    for i, d in enumerate(all_class_dates):
        if d >= ref_date:
            current_index = i
            break

    if all_class_dates[-1] < ref_date:
        current_index = len(all_class_dates) - 1

    # 5. Формируем окно
    start_idx = max(0, current_index - 4)
    end_idx = min(len(all_class_dates), current_index + 6)
    window_dates = all_class_dates[start_idx:end_idx]

    # 6. Подготовка данных о днях
    week_data = []
    for d in window_dates:
        slots_today = ScheduleSlot.objects.filter(group=group, weekday=d.weekday())
        start_time = slots_today.first().start_time if slots_today else None
        week_data.append({
            'date': d,
            'start_time': start_time,
            'is_today': d == today
        })

    # 7. ПОЛУЧАЕМ ДЕТЕЙ (с учетом архива)
    if show_archived:
        children_qs = Child.objects.filter(
            group=group,
            status__in=[Child.Status.ARCHIVED, Child.Status.LOST]
        )
    else:
        children_qs = Child.objects.filter(
            group=group,
            status__in=[Child.Status.ACTIVE, Child.Status.TRIAL]
        )

    children_qs = children_qs.select_related('group__trainer').prefetch_related(
        'subscriptions', 'payments', 'attendances', 'ranks'
    )

    # 8. СОРТИРОВКА (превращаем в список и сортируем)
    children_list = list(children_qs)
    
    if sort_by == 'sessions':
        children_list.sort(key=lambda c: c.sessions_left(), reverse=True)
    elif sort_by == 'debt':
        children_list.sort(key=lambda c: c.debt(), reverse=True)
    else:
        children_list.sort(key=lambda c: (c.last_name.lower(), c.first_name.lower()))

    children_data = []
    for child in children_list:
        active_sub = child.active_subscription()
        projected_end = child.projected_end_date()

        att_map = {}
        for att in child.attendances.filter(date__in=window_dates):
            att_map[att.date] = att.status

        entries = []
        sub_end_index = None
        subscription_ending_soon = False

        for idx, wd in enumerate(week_data):
            status = att_map.get(wd['date'], '')
            entries.append({'date': wd['date'], 'status': status})
            
            if projected_end and wd['date'] == projected_end:
                sub_end_index = idx
            
            if projected_end and wd['date'] >= projected_end - timedelta(days=7) and wd['date'] <= projected_end:
                subscription_ending_soon = True

        # Авто-перевод в потерянные, если пробный истек (только для не-архивных)
        if not show_archived and child.is_trial_expired():
            child.mark_as_lost()

        children_data.append({
            'child': child,
            'initials': f"{child.last_name[0]}{child.first_name[0]}".upper(),
            'age': child.age_display(),
            'sessions_left': child.sessions_left(),
            'sessions_total': active_sub.sessions_total if active_sub else 0,
            'debt': child.debt(),
            'has_certificate': child.has_certificate(),
            'discount_percent': child.discount_percent,
            'subscription_end': active_sub.end_date if active_sub else None,
            'subscription_end_index': sub_end_index,
            'subscription_ending_soon': subscription_ending_soon,
            'is_trial': child.status == Child.Status.TRIAL,
            'is_archived': child.status in [Child.Status.ARCHIVED, Child.Status.LOST],
            'attendance_entries': entries,
        })

    # 9. Переключатель дат
    if len(window_dates) >= 2:
        window_span = (window_dates[-1] - window_dates[0]).days
    else:
        window_span = 7

    prev_ref = (window_dates[0] - timedelta(days=window_span)).strftime('%Y-%m-%d') if window_dates else (today - timedelta(days=7)).strftime('%Y-%m-%d')
    next_ref = (window_dates[-1] + timedelta(days=window_span)).strftime('%Y-%m-%d') if window_dates else (today + timedelta(days=7)).strftime('%Y-%m-%d')

    # Базовые параметры для ссылок, чтобы сортировка не слетала
    base_params = f"group_id={group.id}&ref_date={{}}&sort={sort_by}"
    if show_archived:
        base_params += "&show_archived=1"

    context = {
        'groups': Group.objects.filter(is_active=True),
        'selected_group': group,
        'trainer': trainer,
        'week_data': week_data,
        'children_data': children_data,
        'ref_date': ref_date,
        'prev_ref': prev_ref,
        'next_ref': next_ref,
        'today': today,
        'sort_by': sort_by,
        'show_archived': show_archived,
        'base_params': base_params,
        'page': 'attendance'
    }
    return render(request, 'crm/attendance.html', context)


@login_required
@require_POST
def archive_child_view(request, child_id):
    """Перевод ребенка в архив"""
    child = get_object_or_404(Child, id=child_id)
    child.archive()
    messages.success(request, f'{child.last_name} {child.first_name} архивирован')
    return redirect('attendance')


@login_required
@require_POST
def restore_child_view(request, child_id):
    """Восстановление из архива"""
    child = get_object_or_404(Child, id=child_id)
    child.restore_from_archive()
    messages.success(request, f'{child.last_name} {child.first_name} восстановлен')
    return redirect('attendance')


@login_required
@require_POST
def add_trial_child_view(request, group_id):
    """Добавление ребенка на пробное занятие"""
    group = get_object_or_404(Group, id=group_id)
    
    if request.method == 'POST':
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        parent_phone = request.POST.get('parent_phone')
        
        if last_name and first_name:
            child = Child.objects.create(
                last_name=last_name,
                first_name=first_name,
                parent_phone=parent_phone,
                group=group,
                status=Child.Status.TRIAL,
                trial_from=timezone.localdate()
            )
            messages.success(request, f'{child.last_name} {child.first_name} добавлен на пробное (14 дней)')
            return redirect('attendance')
    
    return redirect('attendance')


@login_required
@require_POST
def cancel_attendance_view(request):
    """Отмена отметки через правый клик"""
    child_id = request.POST.get('child_id')
    date_str = request.POST.get('date')
    
    if child_id and date_str:
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            attendance = Attendance.objects.filter(child_id=child_id, date=date).first()
            if attendance:
                attendance.delete()
                return JsonResponse({'status': 'ok', 'message': 'Отметка отменена'})
        except ValueError:
            pass
    
    return JsonResponse({'status': 'error', 'message': 'Ошибка'}, status=400)

@login_required
@require_POST
def mark_attendance_view(request):
    """Поставить или изменить отметку посещения."""

    child_id = request.POST.get("child_id")
    date_str = request.POST.get("date")
    status = request.POST.get("status")

    if not child_id or not date_str or not status:
        return JsonResponse(
            {
                "status": "error",
                "message": "Не передан child_id, date или status",
            },
            status=400,
        )

    try:
        mark_date = date.fromisoformat(date_str)
    except ValueError:
        return JsonResponse(
            {
                "status": "error",
                "message": "Некорректная дата",
            },
            status=400,
        )

    child = get_object_or_404(Child, pk=child_id)

    allowed_statuses = {
        Attendance.Status.PRESENT,
        Attendance.Status.ABSENT,
        Attendance.Status.EXCUSED,
        Attendance.Status.FROZEN,
        Attendance.Status.VACATION,
    }

    if status not in allowed_statuses:
        return JsonResponse(
            {
                "status": "error",
                "message": "Некорректный статус",
            },
            status=400,
        )

    charge = Decimal("0")

    if (
        status == Attendance.Status.PRESENT
        and not child.active_subscription()
        and child.group
    ):
        charge = child.group.single_session_price

    attendance, created = Attendance.objects.update_or_create(
        child=child,
        date=mark_date,
        slot=None,
        defaults={
            "status": status,
            "charge_amount": charge,
        },
    )

    return JsonResponse(
        {
            "status": "ok",
            "created": created,
            "attendance_id": attendance.pk,
        }
    )

@login_required
def attendance_page(request):
    groups = Group.objects.filter(is_active=True).select_related("trainer")
    selected_group = groups.filter(pk=request.GET.get("group")).first() or groups.first()
    start_raw = request.GET.get("week")
    try:
        chosen = date.fromisoformat(start_raw) if start_raw else timezone.localdate()
    except ValueError:
        chosen = timezone.localdate()
    week_start = chosen - timedelta(days=chosen.weekday())
    week_dates = [week_start + timedelta(days=i) for i in range(14)]

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "mark":
            child = get_object_or_404(Child, pk=request.POST.get("child_id"))
            mark_date = date.fromisoformat(request.POST["date"])
            status = request.POST.get("status", "")
            if status:
                charge = Decimal("0")
                if status == Attendance.Status.PRESENT and not child.active_subscription() and child.group:
                    charge = child.group.single_session_price
                attendance, _ = Attendance.objects.update_or_create(
                    child=child, date=mark_date, slot=None,
                    defaults={"status": status, "charge_amount": charge},
                )
                log_action(request, "attendance.update", attendance, f"Изменена отметка: {child} — {mark_date:%d.%m.%Y}")
            else:
                Attendance.objects.filter(child=child, date=mark_date, slot=None).delete()
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"ok": True})
            return redirect(f"{reverse('attendance')}?group={child.group_id}&week={week_start.isoformat()}")
        if action == "create_child":
            full_name = request.POST.get("full_name", "").split()
            if len(full_name) < 2:
                messages.error(request, "Укажите фамилию и имя ребёнка")
            else:
                child = Child.objects.create(
                    last_name=full_name[0],
                    first_name=full_name[1],
                    patronymic=" ".join(full_name[2:]),
                    birth_year=int(request.POST.get("birth_year") or timezone.localdate().year - 7),
                    parent_phone=request.POST.get("parent_phone", ""),
                    parent_name=request.POST.get("parent_name", ""),
                    group_id=request.POST.get("group_id") or None,
                    status=Child.Status.TRIAL if request.POST.get("trial") else Child.Status.ACTIVE,
                    trial_from=timezone.localdate() if request.POST.get("trial") else None,
                )
                log_action(request, "child.create", child, f"Добавлен спортсмен {child}")
                messages.success(request, "Спортсмен добавлен")
            return redirect("attendance")

    children = []
    if selected_group:
        queryset = selected_group.children.filter(status__in=[Child.Status.ACTIVE, Child.Status.TRIAL])
        search = request.GET.get("q", "").strip()
        if search:
            queryset = queryset.filter(Q(last_name__icontains=search) | Q(first_name__icontains=search) | Q(parent_phone__icontains=search))
        marks = Attendance.objects.filter(child__in=queryset, date__range=(week_dates[0], week_dates[-1]))
        mark_map = {(m.child_id, m.date): m.status for m in marks}
        style_map = {
            Attendance.Status.PRESENT: ("+", "bg-emerald-100 text-emerald-700"),
            Attendance.Status.ABSENT: ("×", "bg-red-100 text-red-700"),
            Attendance.Status.EXCUSED: ("У", "bg-purple-100 text-purple-700"),
            Attendance.Status.FROZEN: ("❄", "bg-blue-100 text-blue-700"),
            Attendance.Status.VACATION: ("О", "bg-amber-100 text-amber-700"),
        }
        for child in queryset:
            sub = child.active_subscription()
            child_marks = []
            for day in week_dates:
                status = mark_map.get((child.pk, day), "")
                symbol, css = style_map.get(status, ("", "bg-slate-100 text-slate-400"))
                child_marks.append({
                    "date": day,
                    "status": status,
                    "symbol": symbol,
                    "css": css,
                    "week_number": 1 if day < week_start + timedelta(days=7) else 2,
                    "is_expiry": child.nearest_expiry() == day,
                })
            children.append({
                "child": child,
                "form": ChildForm(instance=child, prefix=f"child-{child.pk}"),
                "subscription": sub,
                "left": child.sessions_left(),
                "debt": child.debt(),
                "expiry": child.nearest_expiry(),
                "marks": child_marks,
                "attendance_history": child.attendances.all()[:8],
                "competition_history": child.competition_entries.select_related("competition")[:8],
            })
        if request.GET.get("sort") == "remaining":
            children.sort(key=lambda item: item["left"])
    return render(request, "crm/attendance.html", page_context(
        request, "attendance", groups=groups, selected_group=selected_group,
        week_dates=week_dates, week_start=week_start, prev_week=week_start-timedelta(days=14),
        next_week=week_start+timedelta(days=14), children=children, today=timezone.localdate(),
    ))


@login_required
def child_edit(request, pk):
    child = get_object_or_404(Child, pk=pk)
    next_url = request.POST.get("next") or reverse("attendance")
    if request.method != "POST":
        return redirect(next_url)
    action = request.POST.get("action", "save")
    if action == "archive":
        child.status = Child.Status.ARCHIVED
        child.save(update_fields=["status"])
        log_action(request, "child.archive", child, f"{child} перемещён в архив")
        messages.success(request, "Спортсмен перемещён в архив")
        return redirect(next_url)
    form = ChildForm(request.POST, request.FILES, instance=child, prefix=f"child-{child.pk}")
    if form.is_valid():
        child = form.save()
        if child.status != Child.Status.TRIAL:
            child.trial_from = None
            child.save(update_fields=["trial_from"])
        log_action(request, "child.update", child, f"Обновлена карточка спортсмена {child}")
        messages.success(request, "Карточка спортсмена обновлена")
    else:
        messages.error(request, "Не удалось сохранить карточку: " + "; ".join(f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()))
    return redirect(next_url)


@login_required
def statistics_page(request):
    today = timezone.localdate()
    month_start = today.replace(day=1)
    active = Child.objects.filter(status=Child.Status.ACTIVE)
    total = Child.objects.exclude(status=Child.Status.ARCHIVED).count()
    new = Child.objects.filter(created_at__date__gte=month_start).count()
    lost = Child.objects.filter(status=Child.Status.LOST).count()
    revenue = Payment.objects.filter(date__gte=month_start).aggregate(value=Sum("amount"))["value"] or 0
    group_stats = Group.objects.filter(is_active=True).annotate(
        children_count=Count("children", filter=Q(children__status=Child.Status.ACTIVE), distinct=True),
        attended=Count("children__attendances", filter=Q(children__attendances__status=Attendance.Status.PRESENT, children__attendances__date__gte=month_start)),
        marked=Count("children__attendances", filter=Q(children__attendances__status__in=[Attendance.Status.PRESENT, Attendance.Status.ABSENT], children__attendances__date__gte=month_start)),
    )
    for group in group_stats:
        group.attendance_percent = round(group.attended * 100 / group.marked) if group.marked else 0
    return render(request, "crm/statistics.html", page_context(
        request, "statistics", total=total, active_count=active.count(), new_count=new,
        lost_count=lost, revenue=revenue, group_stats=group_stats,
    ))


@login_required
def payments_page(request):
    editing_tariff = Tariff.objects.filter(pk=request.GET.get("edit_tariff")).first()
    editing_subscription = Subscription.objects.filter(pk=request.GET.get("edit_subscription")).first()
    tariff_form = TariffForm(request.POST or None, prefix="tariff", instance=editing_tariff)
    subscription_form = SubscriptionForm(
        request.POST or None, prefix="subscription", instance=editing_subscription,
        initial={"start_date": timezone.localdate()},
    )
    if request.method == "POST":
        action = request.POST.get("action", "payment")
        if action == "payment":
            child = get_object_or_404(Child, pk=request.POST.get("child_id"))
            try:
                amount = Decimal(request.POST.get("amount", "0"))
            except InvalidOperation:
                amount = Decimal("0")
            if amount <= 0:
                messages.error(request, "Сумма должна быть больше нуля")
            else:
                payment = Payment.objects.create(
                    child=child,
                    subscription=child.active_subscription(),
                    amount=amount,
                    date=request.POST.get("date") or timezone.localdate(),
                    created_by=request.user,
                )
                log_action(request, "payment.create", payment, f"Принята оплата {amount} ₽ от {child}")
                messages.success(request, "Оплата сохранена")
        elif action == "save_tariff":
            tariff_form = TariffForm(request.POST, prefix="tariff", instance=editing_tariff)
            if tariff_form.is_valid():
                tariff = tariff_form.save()
                log_action(request, "tariff.save", tariff, f"Сохранён тариф {tariff.name}")
                messages.success(request, "Тариф сохранён")
            else:
                messages.error(request, "Проверьте параметры тарифа")
                return render(request, "crm/payments.html", page_context(
                    request, "payments", tariff_form=tariff_form,
                    subscription_form=SubscriptionForm(prefix="subscription"), tariffs=Tariff.objects.all(),
                    subscriptions=Subscription.objects.select_related("child", "tariff"), children=Child.objects.filter(status=Child.Status.ACTIVE),
                    renewals=[], urgent_count=0, expected=0, editing_tariff=editing_tariff,
                ))
        elif action == "toggle_tariff":
            tariff = get_object_or_404(Tariff, pk=request.POST.get("tariff_id"))
            tariff.is_active = not tariff.is_active
            tariff.save(update_fields=["is_active"])
            messages.success(request, "Статус тарифа изменён")
        elif action == "save_subscription":
            subscription_form = SubscriptionForm(request.POST, prefix="subscription", instance=editing_subscription)
            if subscription_form.is_valid():
                subscription = subscription_form.save()
                Subscription.objects.filter(
                    child=subscription.child, is_active=True,
                ).exclude(pk=subscription.pk).update(is_active=False)
                if subscription.child.status == Child.Status.TRIAL:
                    subscription.child.status = Child.Status.ACTIVE
                    subscription.child.trial_from = None
                    subscription.child.save(update_fields=["status", "trial_from"])
                log_action(request, "subscription.save", subscription, f"Сохранён абонемент {subscription}")
                messages.success(request, "Абонемент назначен")
            else:
                messages.error(request, "Проверьте данные абонемента")
                return redirect("payments")
        elif action == "cancel_subscription":
            subscription = get_object_or_404(Subscription, pk=request.POST.get("subscription_id"))
            subscription.is_active = False
            subscription.save(update_fields=["is_active"])
            log_action(request, "subscription.cancel", subscription, f"Отменён абонемент {subscription}")
            messages.success(request, "Абонемент отменён без удаления истории")
        return redirect("payments")
    today = timezone.localdate()
    limit = today + timedelta(days=14)
    renewals = []
    for child in Child.objects.filter(status=Child.Status.ACTIVE).select_related("group"):
        sub = child.active_subscription()
        if sub and sub.end_date <= limit:
            renewals.append({"child": child, "subscription": sub, "debt": child.debt(), "days": (sub.end_date-today).days})
    renewals.sort(key=lambda item: item["subscription"].end_date)
    return render(request, "crm/payments.html", page_context(
        request, "payments", renewals=renewals, children=Child.objects.filter(status=Child.Status.ACTIVE),
        urgent_count=sum(1 for x in renewals if x["days"] <= 3), expected=sum((x["subscription"].price for x in renewals), Decimal("0")),
        tariffs=Tariff.objects.all(), subscriptions=Subscription.objects.select_related("child", "tariff")[:100],
        tariff_form=tariff_form, subscription_form=subscription_form,
        editing_tariff=editing_tariff, editing_subscription=editing_subscription,
    ))


@login_required
def expenses_page(request):
    editing = Expense.objects.filter(pk=request.GET.get("edit")).first()
    form = ExpenseForm(request.POST or None, request.FILES or None, instance=editing)
    if request.method == "POST":
        if request.POST.get("action") == "delete":
            expense = get_object_or_404(Expense, pk=request.POST.get("expense_id"))
            log_action(request, "expense.delete", expense, f"Удалён расход {expense.title} на {expense.amount} ₽")
            expense.delete()
            messages.success(request, "Расход удалён")
            return redirect("expenses")
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            log_action(request, "expense.save", expense, f"Сохранён расход {expense.title} на {expense.amount} ₽")
            messages.success(request, "Расход сохранён")
            return redirect("expenses")
        messages.error(request, "Проверьте заполнение формы")
    today = timezone.localdate()
    month_start = today.replace(day=1)
    expenses = Expense.objects.select_related("created_by")
    month_expenses = expenses.filter(date__gte=month_start)
    total = month_expenses.aggregate(value=Sum("amount"))["value"] or 0
    household = month_expenses.filter(category=Expense.Category.HOUSEHOLD).aggregate(value=Sum("amount"))["value"] or 0
    return render(request, "crm/expenses.html", page_context(
        request, "expenses", form=form, expenses=expenses, editing=editing,
        total=total, household=household, operations=month_expenses.count(),
    ))


def recalculate_places(competition):
    for category in competition.entries.values_list("category", flat=True).distinct():
        entries = list(competition.entries.filter(category=category))
        entries.sort(key=lambda entry: entry.total_points(), reverse=True)
        last_total = None
        last_place = 0
        for index, entry in enumerate(entries, 1):
            total = entry.total_points()
            place = last_place if total == last_total else index
            CompetitionEntry.objects.filter(pk=entry.pk).update(place=place)
            last_total, last_place = total, place


@login_required
def competitions_page(request):
    competitions = Competition.objects.all()
    selected = competitions.filter(pk=request.GET.get("competition")).first() or competitions.first()
    editing_competition = competitions.filter(pk=request.GET.get("edit_competition")).first()
    editing_apparatus = Apparatus.objects.filter(pk=request.GET.get("edit_apparatus"), competition=selected).first() if selected else None
    competition_form = CompetitionForm(prefix="competition", instance=editing_competition)
    entry_form = CompetitionEntryForm(prefix="entry")
    apparatus_form = ApparatusForm(prefix="apparatus", instance=editing_apparatus)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_competition":
            target_competition = competitions.filter(pk=request.POST.get("competition_id")).first()
            competition_form = CompetitionForm(request.POST, prefix="competition", instance=target_competition)
            if competition_form.is_valid():
                selected = competition_form.save()
                if not target_competition:
                    for order, name in enumerate(("Прыжок", "Брусья", "Бревно", "Вольные")):
                        Apparatus.objects.create(competition=selected, name=name, order=order)
                log_action(request, "competition.save", selected, f"Сохранено соревнование {selected.name}")
                messages.success(request, "Соревнование сохранено")
                return redirect(f"{reverse('competitions')}?competition={selected.pk}")
        elif action == "save_apparatus" and selected:
            target_apparatus = selected.apparatus.filter(pk=request.POST.get("apparatus_id")).first()
            apparatus_form = ApparatusForm(request.POST, prefix="apparatus", instance=target_apparatus)
            if apparatus_form.is_valid():
                apparatus_item = apparatus_form.save(commit=False)
                apparatus_item.competition = selected
                apparatus_item.save()
                for entry in selected.entries.all():
                    ApparatusScore.objects.get_or_create(entry=entry, apparatus=apparatus_item, defaults={"points": 0})
                log_action(request, "competition.apparatus", apparatus_item, f"Сохранена дисциплина {apparatus_item.name}")
                messages.success(request, "Дисциплина сохранена")
                return redirect(f"{reverse('competitions')}?competition={selected.pk}")
        elif action == "delete_apparatus" and selected:
            apparatus_item = get_object_or_404(selected.apparatus, pk=request.POST.get("apparatus_id"))
            description = apparatus_item.name
            apparatus_item.delete()
            log_action(request, "competition.apparatus.delete", selected, f"Удалена дисциплина {description}")
            messages.success(request, "Дисциплина и её баллы удалены")
            return redirect(f"{reverse('competitions')}?competition={selected.pk}")
        elif action == "add_entry" and selected:
            entry_form = CompetitionEntryForm(request.POST, prefix="entry")
            if entry_form.is_valid():
                entry = entry_form.save(commit=False)
                entry.competition = selected
                entry.save()
                for apparatus in selected.apparatus.all():
                    ApparatusScore.objects.create(entry=entry, apparatus=apparatus, points=0)
                log_action(request, "competition.entry", entry, f"Добавлен участник {entry.child} в {selected.name}")
                messages.success(request, "Участник добавлен")
                return redirect(f"{reverse('competitions')}?competition={selected.pk}")
        elif action == "save_scores" and selected:
            for entry in selected.entries.all():
                for apparatus in selected.apparatus.all():
                    key = f"score_{entry.pk}_{apparatus.pk}"
                    try:
                        points = Decimal(request.POST.get(key, "0").replace(",", "."))
                    except InvalidOperation:
                        points = Decimal("0")
                    ApparatusScore.objects.update_or_create(entry=entry, apparatus=apparatus, defaults={"points": points})
            recalculate_places(selected)
            log_action(request, "competition.scores", selected, f"Обновлены результаты {selected.name}")
            messages.success(request, "Баллы сохранены, места пересчитаны")
            return redirect(f"{reverse('competitions')}?competition={selected.pk}")
        messages.error(request, "Проверьте заполнение формы")
    entry_rows = []
    apparatus = list(selected.apparatus.all()) if selected else []
    if selected:
        for entry in selected.entries.select_related("child").prefetch_related("scores"):
            score_map = {score.apparatus_id: score for score in entry.scores.all()}
            entry_rows.append({
                "entry": entry,
                "scores": [
                    {"apparatus_id": item.pk, "points": score_map[item.pk].points if item.pk in score_map else Decimal("0")}
                    for item in apparatus
                ],
                "total": entry.total_points(),
            })
    return render(request, "crm/competitions.html", page_context(
        request, "competitions", competitions=competitions, selected=selected,
        apparatus=apparatus, entry_rows=entry_rows, competition_form=competition_form,
        entry_form=entry_form, apparatus_form=apparatus_form,
        editing_competition=editing_competition, editing_apparatus=editing_apparatus,
    ))


@login_required
def competition_export(request, pk):
    competition = get_object_or_404(Competition, pk=pk)
    apparatus = list(competition.apparatus.all())
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    headers = ["Спортсмен", "Год рождения", "Разряд", "Категория", *[a.name for a in apparatus], "Итого", "Место"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17202A")
        cell.alignment = Alignment(horizontal="center")
    for entry in competition.entries.select_related("child").prefetch_related("scores"):
        score_map = {score.apparatus_id: score.points for score in entry.scores.all()}
        ws.append([str(entry.child), entry.child.birth_year, entry.rank, entry.category, *[float(score_map.get(a.pk, 0)) for a in apparatus], float(entry.total_points()), entry.place or ""])
    widths = [28, 14, 14, 18, *([12] * len(apparatus)), 12, 10]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="competition-{competition.pk}.xlsx"'
    wb.save(response)
    log_action(request, "competition.export", competition, f"Выгружены результаты {competition.name}")
    return response


@login_required
def notifications_page(request):
    if request.method == "POST":
        task = get_object_or_404(ManagerTask, pk=request.POST.get("task_id"))
        if task.assignee_id not in (None, request.user.pk) and user_role(request.user) != Role.BOSS:
            return HttpResponseForbidden("Это задача другого пользователя")
        task.is_done = not task.is_done
        task.done_at = timezone.now() if task.is_done else None
        task.save(update_fields=["is_done", "done_at"])
        log_action(request, "task.toggle", task, f"Задача «{task.title}»: {'выполнена' if task.is_done else 'возвращена'}")
        messages.success(request, "Статус задачи изменён")
        return redirect("notifications")
    tasks = ManagerTask.objects.select_related("assignee", "created_by")
    if user_role(request.user) != Role.BOSS:
        tasks = tasks.filter(Q(assignee=request.user) | Q(assignee__isnull=True))
    today = timezone.localdate()
    alerts = []
    for child in Child.objects.filter(status=Child.Status.ACTIVE):
        expiry = child.nearest_expiry()
        debt = child.debt()
        if expiry and expiry <= today + timedelta(days=7):
            alerts.append({"kind": "subscription", "child": child, "expiry": expiry, "debt": debt})
    trials = Child.objects.filter(status=Child.Status.TRIAL, trial_from__lte=today)
    reminders = Reminder.objects.filter(
        Q(assignee=request.user) | Q(visible_to_all=True),
        is_done=False,
        remind_at__date__lte=today + timedelta(days=1),
    ).select_related("assignee")
    imported_leads = Lead.objects.filter(imported_from_ad=True, status=Lead.Status.NEW)
    return render(request, "crm/notifications.html", page_context(
        request, "notifications", tasks=tasks, alerts=alerts, trials=trials,
        reminders=reminders, imported_leads=imported_leads,
        open_count=tasks.filter(is_done=False).count() + len(alerts) + trials.count() + reminders.count() + imported_leads.count(), today=today,
    ))


@login_required
def applications_page(request):
    editing = Lead.objects.filter(pk=request.GET.get("edit")).first()
    form = LeadForm(request.POST or None, instance=editing)
    if request.method == "POST":
        action = request.POST.get("action", "save")
        if action == "import_raw":
            raw = request.POST.get("raw_application", "")
            parsed = {}
            aliases = {"имя": "full_name", "фио": "full_name", "телефон": "phone", "возраст": "age_text", "источник": "source", "кампания": "source", "комментарий": "comment"}
            for line in raw.splitlines():
                if ":" not in line:
                    continue
                key, value = line.split(":", 1)
                target = aliases.get(key.strip().lower())
                if target and value.strip():
                    parsed[target] = value.strip()
            if parsed.get("full_name"):
                parsed.setdefault("source", "VK Реклама")
                lead = Lead.objects.create(imported_from_ad=True, **parsed)
                log_action(request, "lead.import", lead, f"Импортирована рекламная заявка {lead.full_name}")
                messages.success(request, "Заявка распознана и выделена как рекламная")
            else:
                messages.error(request, "Не удалось распознать имя. Используйте строку «Имя: ...»")
            return redirect("applications")
        if action == "create_newcomer":
            lead = get_object_or_404(Lead, pk=request.POST.get("lead_id"))
            newcomer = Newcomer.objects.create(
                lead=lead,
                full_name=lead.full_name,
                birth_date=lead.birth_date,
                age_text=lead.age_text,
                phone=lead.phone,
                source=lead.source,
                trial_at=lead.trial_at,
                trainer=lead.trainer,
                group=lead.group,
                comment=lead.comment,
            )
            lead.status = Lead.Status.QUALIFIED
            lead.save(update_fields=["status"])
            log_action(request, "lead.newcomer", newcomer, f"Из заявки создан новичок {newcomer.full_name}")
            messages.success(request, "Новичок создан и предзаполнен из заявки")
            return redirect("newcomers")
        if form.is_valid():
            lead = form.save()
            log_action(request, "lead.save", lead, f"Сохранена заявка {lead.full_name}")
            messages.success(request, "Заявка сохранена")
            return redirect("applications")
        messages.error(request, "Проверьте данные заявки")
    leads = Lead.objects.select_related("trainer", "group").prefetch_related("newcomers")
    return render(request, "crm/applications.html", page_context(
        request, "applications", leads=leads, form=form, editing=editing,
        imported_count=leads.filter(imported_from_ad=True, status=Lead.Status.NEW).count(),
    ))


@login_required
def newcomers_page(request):
    editing = Newcomer.objects.filter(pk=request.GET.get("edit")).first()
    form = NewcomerForm(request.POST or None, instance=editing)
    if request.method == "POST":
        action = request.POST.get("action", "save")
        if action == "convert":
            newcomer = get_object_or_404(Newcomer, pk=request.POST.get("newcomer_id"))
            if newcomer.child:
                messages.info(request, "Карточка спортсмена уже создана")
                return redirect("newcomers")
            parts = newcomer.full_name.split()
            child = Child.objects.create(
                last_name=parts[0] if parts else "Без фамилии",
                first_name=parts[1] if len(parts) > 1 else "Без имени",
                patronymic=" ".join(parts[2:]),
                birth_date=newcomer.birth_date,
                birth_year=newcomer.birth_date.year if newcomer.birth_date else timezone.localdate().year - 7,
                parent_phone=newcomer.phone,
                group=newcomer.group,
                status=Child.Status.ACTIVE if newcomer.paid else Child.Status.TRIAL,
                trial_from=None if newcomer.paid else timezone.localdate(),
                note=newcomer.comment,
            )
            newcomer.child = child
            newcomer.save(update_fields=["child"])
            if newcomer.lead:
                newcomer.lead.child = child
                newcomer.lead.save(update_fields=["child"])
            log_action(request, "newcomer.convert", child, f"Новичок {newcomer.full_name} перенесён в спортсмены")
            messages.success(request, "Карточка спортсмена создана")
            return redirect("payments" if newcomer.paid else "attendance")
        if form.is_valid():
            newcomer = form.save()
            log_action(request, "newcomer.save", newcomer, f"Сохранён новичок {newcomer.full_name}")
            messages.success(request, "Новичок сохранён")
            return redirect("newcomers")
        messages.error(request, "Проверьте данные новичка")
    return render(request, "crm/newcomers.html", page_context(
        request, "newcomers", newcomers=Newcomer.objects.select_related("lead", "trainer", "group", "child"),
        form=form, editing=editing,
    ))


@login_required
def calendar_page(request):
    editing = Reminder.objects.filter(pk=request.GET.get("edit")).first()
    if editing and editing.created_by_id != request.user.pk and user_role(request.user) != Role.BOSS:
        return HttpResponseForbidden("Редактировать чужое напоминание может только начальник")
    form = ReminderForm(request.POST or None, instance=editing, initial={"assignee": request.user})
    if request.method == "POST":
        action = request.POST.get("action", "save")
        if action == "set_shift":
            profile, _ = StaffProfile.objects.get_or_create(user=request.user)
            try:
                profile.shift_anchor = date.fromisoformat(request.POST.get("shift_anchor", ""))
                profile.save(update_fields=["shift_anchor"])
                messages.success(request, "График 2/2 пересчитан")
            except ValueError:
                messages.error(request, "Укажите первый рабочий день")
            return redirect("calendar")
        if action == "toggle":
            reminder = get_object_or_404(Reminder, pk=request.POST.get("reminder_id"))
            if reminder.assignee_id != request.user.pk and user_role(request.user) != Role.BOSS:
                return HttpResponseForbidden("Недостаточно прав")
            reminder.is_done = not reminder.is_done
            reminder.save(update_fields=["is_done"])
            return redirect("calendar")
        if form.is_valid():
            reminder = form.save(commit=False)
            reminder.created_by = editing.created_by if editing else request.user
            reminder.save()
            log_action(request, "reminder.save", reminder, f"Сохранено напоминание {reminder.title}")
            messages.success(request, "Напоминание сохранено")
            return redirect("calendar")
        messages.error(request, "Проверьте дату и поля напоминания")
    visible = Reminder.objects.select_related("assignee", "created_by").filter(
        Q(assignee=request.user) | Q(created_by=request.user) | Q(visible_to_all=True)
    ).distinct()
    if user_role(request.user) == Role.BOSS:
        visible = Reminder.objects.select_related("assignee", "created_by")
    week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
    days = []
    for offset in range(14):
        day = week_start + timedelta(days=offset)
        day_items = [item for item in visible if timezone.localtime(item.remind_at).date() == day]
        days.append({"date": day, "items": day_items})
    shift_rows = []
    for profile in StaffProfile.objects.select_related("user", "branch").filter(user__is_active=True).exclude(role=Role.BOSS):
        work_days = set()
        if profile.shift_anchor:
            for item in days:
                if (item["date"] - profile.shift_anchor).days % 4 in (0, 1):
                    work_days.add(item["date"])
        shift_rows.append({"profile": profile, "work_days": work_days})
    return render(request, "crm/calendar.html", page_context(
        request, "calendar", form=form, editing=editing, days=days, reminders=visible,
        today=timezone.localdate(), shift_rows=shift_rows,
    ))


@login_required
def search_page(request):
    query = request.GET.get("q", "").strip()
    children = leads = groups = []
    if query:
        children = Child.objects.filter(
            Q(last_name__icontains=query) | Q(first_name__icontains=query) |
            Q(parent_name__icontains=query) | Q(parent_phone__icontains=query)
        ).select_related("group")[:30]
        leads = Lead.objects.filter(Q(full_name__icontains=query) | Q(phone__icontains=query))[:30]
        groups = Group.objects.filter(Q(name__icontains=query) | Q(trainer__full_name__icontains=query))[:20]
    return render(request, "crm/search.html", page_context(
        request, "search", query=query, children=children, leads=leads, groups=groups,
    ))


@role_required(1)
def backup_export(request):
    output = StringIO()
    call_command("dumpdata", "crm", indent=2, stdout=output)
    response = HttpResponse(output.getvalue(), content_type="application/json; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="crm-backup-{timezone.localdate():%Y-%m-%d}.json"'
    log_action(request, "backup.export", None, "Выгружена резервная копия данных CRM")
    return response


@role_required(2)
def boss_page(request):
    today = timezone.localdate()
    month = today.replace(day=1)
    target = RevenueTarget.objects.filter(month=month).first()
    task_form = ManagerTaskForm(prefix="task")
    target_form = RevenueTargetForm(prefix="target", instance=target, initial={"month": month})
    if request.method == "POST":
        if request.POST.get("action") == "create_task":
            task_form = ManagerTaskForm(request.POST, prefix="task")
            if task_form.is_valid():
                task = task_form.save(commit=False)
                task.created_by = request.user
                task.save()
                log_action(request, "task.create", task, f"Поставлена задача «{task.title}»")
                messages.success(request, "Задача поставлена")
                return redirect("boss")
        elif request.POST.get("action") == "set_target":
            target_form = RevenueTargetForm(request.POST, prefix="target", instance=target)
            if target_form.is_valid():
                target = target_form.save(commit=False)
                target.set_by = request.user
                target.save()
                log_action(request, "revenue_target.save", target, f"Цель выручки: {target.amount} ₽")
                messages.success(request, "Цель обновлена")
                return redirect("boss")
        messages.error(request, "Проверьте заполнение формы")
    revenue = Payment.objects.filter(date__gte=month).aggregate(value=Sum("amount"))["value"] or 0
    target_amount = target.amount if target else Decimal("0")
    target_percent = min(100, round(revenue * 100 / target_amount)) if target_amount else 0
    salaries = SalaryPayout.objects.filter(month=month).aggregate(value=Sum("amount"))["value"] or 0
    trainer_rows = []
    for trainer in Trainer.objects.filter(is_active=True):
        children = Child.objects.filter(group__trainer=trainer, status=Child.Status.ACTIVE)
        marked = Attendance.objects.filter(child__in=children, date__gte=month, status__in=[Attendance.Status.PRESENT, Attendance.Status.ABSENT]).count()
        present = Attendance.objects.filter(child__in=children, date__gte=month, status=Attendance.Status.PRESENT).count()
        trial = Child.objects.filter(group__trainer=trainer, status=Child.Status.TRIAL).count()
        lost = Child.objects.filter(group__trainer=trainer, status=Child.Status.LOST).count()
        trainer_rows.append({"trainer": trainer, "children": children.count(), "trial": trial, "lost": lost, "attendance": round(present*100/marked) if marked else 0})
    trainer_rows.sort(key=lambda row: row["attendance"], reverse=True)
    all_logs = request.GET.get("all_logs") == "1"
    events = AuditEvent.objects.select_related("actor")
    if not all_logs:
        events = events[:12]
    return render(request, "crm/boss.html", page_context(
        request, "boss", revenue=revenue, target=target, target_amount=target_amount,
        target_percent=target_percent, salaries=salaries, active_count=Child.objects.filter(status=Child.Status.ACTIVE).count(),
        trainer_rows=trainer_rows, events=events, all_logs=all_logs,
        overdue_tasks=ManagerTask.objects.filter(is_done=False, due_date__lt=today).count(),
        task_form=task_form, target_form=target_form,
    ))


@role_required(1)
def users_page(request):
    form = StaffCreateForm(request.POST or None)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create" and form.is_valid():
            if form.cleaned_data["role"] == Role.BOSS and user_role(request.user) != Role.BOSS:
                form.add_error("role", "Только начальник может создать аккаунт начальника")
            else:
                user = form.save()
                log_action(request, "user.create", user, f"Создан пользователь {user.username}")
                messages.success(request, "Пользователь создан")
                return redirect("users")
        if action == "toggle":
            user = get_object_or_404(get_user_model(), pk=request.POST.get("user_id"))
            if user == request.user:
                messages.error(request, "Нельзя отключить собственный аккаунт")
            elif user_role(user) == Role.BOSS and user_role(request.user) != Role.BOSS:
                return HttpResponseForbidden("Только начальник управляет аккаунтом начальника")
            else:
                user.is_active = not user.is_active
                user.save(update_fields=["is_active"])
                log_action(request, "user.toggle", user, f"Аккаунт {user.username}: {'включён' if user.is_active else 'отключён'}")
                messages.success(request, "Статус пользователя изменён")
            return redirect("users")
        messages.error(request, "Проверьте форму пользователя")
    users = get_user_model().objects.select_related("profile").filter(is_staff=True).order_by("-is_active", "last_name", "username")
    return render(request, "crm/users.html", page_context(request, "users", users=users, form=form))


@login_required
def profile_page(request):
    profile_form = ProfileForm(instance=request.user, prefix="profile")
    password_form = StyledPasswordChangeForm(request.user, prefix="password")
    if request.method == "POST":
        if request.POST.get("action") == "profile":
            profile_form = ProfileForm(request.POST, instance=request.user, prefix="profile")
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, "Профиль обновлён")
                return redirect("profile")
        elif request.POST.get("action") == "password":
            password_form = StyledPasswordChangeForm(request.user, request.POST, prefix="password")
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                log_action(request, "password.change", user, "Пользователь сменил пароль")
                messages.success(request, "Пароль изменён")
                return redirect("profile")
        messages.error(request, "Проверьте введённые данные")
    return render(request, "crm/profile.html", page_context(
        request, "profile", profile_form=profile_form, password_form=password_form,
    ))


from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .models import Child, Group

# views.py

from .models import Child, Group, calculate_projected_end_date # Не забудь импорт функции!

@login_required
def revenue_forecast_view(request):
    today = timezone.localdate()
    days = []
    processed_child_ids = set()

    for i in range(4):
        current_date = today + timedelta(days=i)
        
        children = Child.objects.filter(status=Child.Status.ACTIVE).select_related('group')
        
        urgent_list = []
        one_left_list = []
        forecast_list = []

        for child in children:
            if child.id in processed_child_ids:
                continue

            active_sub = child.active_subscription()
            if not active_sub:
                continue

            left_on_date = child.sessions_left_on_date(current_date)

            # Блок 1: Срочно (0 занятий)
            if left_on_date == 0:
                urgent_list.append({
                    'name': f"{child.last_name} {child.first_name}",
                    'parent_name': child.parent_name,
                    'parent_phone': child.parent_phone,
                    'amount': active_sub.price
                })
                processed_child_ids.add(child.id)
                continue

            # Блок 2: Осталось 1 занятие
            if left_on_date == 1:
                one_left_list.append({
                    'name': f"{child.last_name} {child.first_name}",
                    'parent_name': child.parent_name,
                    'parent_phone': child.parent_phone,
                    'amount': active_sub.price
                })
                processed_child_ids.add(child.id)
                continue

            # Блок 3: Прогноз (до окончания менее 5 календарных дней)
            proj_end = calculate_projected_end_date(child.group, current_date, left_on_date)
            if proj_end and (proj_end - current_date).days < 5:
                forecast_list.append({
                    'name': f"{child.last_name} {child.first_name}",
                    'parent_name': child.parent_name,
                    'parent_phone': child.parent_phone,
                    'amount': active_sub.price
                })
                processed_child_ids.add(child.id)

        total_day = sum(item['amount'] for item in urgent_list + one_left_list + forecast_list)

        days.append({
            'date': current_date,
            'weekday': current_date.strftime("%A"),
            'total': total_day,
            'urgent': urgent_list,
            'one_left': one_left_list,
            'forecast': forecast_list,
        })

    context = {
        'days': days,
        'title': 'Прогноз доходов',
        'subtitle': 'Ожидаемые продления на ближайшие 4 дня',
        'page': 'revenue_forecast'
    }
    return render(request, 'crm/revenue_forecast.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from .models import Child, Group, ScheduleSlot, Subscription, Attendance, ChildRank
from .forms import ChildForm, SubscriptionForm  # создадим ниже

@login_required
def child_card_view(request, child_id):
    child = get_object_or_404(Child, id=child_id)

    subscriptions = child.subscriptions.all().order_by('-start_date')
    attendances = child.attendances.all().order_by('-date')[:50]
    ranks = child.ranks.all().order_by('-year')
    competitions = child.competition_entries.all().select_related('competition').order_by('-competition__date')[:20]
    camps = child.camp_stays.all().select_related('camp').order_by('-start_date')

    active_sub = child.active_subscription()
    sessions_left = child.sessions_left()
    debt = child.debt()
    balance = child.balance()
    missed_pct = child.missed_percent()
    nearest_exp = child.nearest_expiry()
    promos = child.active_promos()
    groups_list = Group.objects.filter(is_active=True)

    # === HEAT-MAP: последние 365 дней ===
        # === HEAT-MAP: последние 180 дней ===
    today = timezone.localdate()
    days_back = 180
    year_ago = today - timedelta(days=days_back - 1)

    # Получаем все посещения за период
    period_attendances = {
        att.date: att.status
        for att in child.attendances.filter(date__gte=year_ago)
    }

    # Начинаем с понедельника (weekday() возвращает 0=Пн, 6=Вс)
    start_date = year_ago - timedelta(days=year_ago.weekday())
    end_date = today

    weeks = []
    current = start_date
    week_index = 0
    while current <= end_date:
        week = []
        week_start_date = current  # Понедельник этой недели
        for day_in_week in range(7):  # 0=Пн ... 6=Вс
            date = current + timedelta(days=day_in_week)
            if date > end_date:
                week.append(None)
            else:
                status = period_attendances.get(date)
                week.append({
                    'date': date,
                    'status': status,
                    'is_future': date > today,
                })
        weeks.append({
            'days': week,
            'week_start': week_start_date,
            'show_label': week_index % 5 == 0,  # Каждый 5-й столбец
        })
        current += timedelta(days=7)
        week_index += 1

    # Статистика по статусам за период
    period_stats = {
        'present': sum(1 for s in period_attendances.values() if s == 'present'),
        'absent': sum(1 for s in period_attendances.values() if s == 'absent'),
        'frozen': sum(1 for s in period_attendances.values() if s == 'frozen'),
        'vacation': sum(1 for s in period_attendances.values() if s == 'vacation'),
        'excused': sum(1 for s in period_attendances.values() if s == 'excused'),
    }

    if request.method == 'POST' and 'change_group' in request.POST:
        new_group_id = request.POST.get('new_group')
        if new_group_id:
            new_group = get_object_or_404(Group, id=new_group_id)
            child.group = new_group
            child.save(update_fields=['group'])
            messages.success(request, f'Ребенок переведен в группу "{new_group.name}"')
            return redirect('child_card', child_id=child.id)

    context = {
        'child': child,
        'subscriptions': subscriptions,
        'attendances': attendances,
        'ranks': ranks,
        'competitions': competitions,
        'camps': camps,
        'active_sub': active_sub,
        'sessions_left': sessions_left,
        'debt': debt,
        'balance': balance,
        'missed_percent': missed_pct,
        'nearest_expiry': nearest_exp,
        'promos': promos,
        'groups_list': groups_list,
        'weeks': weeks,
        'period_stats': period_stats,
        'today': today,
        'page': 'child_card'
    }
    return render(request, 'crm/child_card.html', context)


@login_required
def child_edit_view(request, child_id):
    child = get_object_or_404(Child, id=child_id)
    if request.method == 'POST':
        form = ChildForm(request.POST, request.FILES, instance=child)
        if form.is_valid():
            form.save()
            messages.success(request, 'Данные ребенка обновлены')
            return redirect('child_card', child_id=child.id)
    else:
        form = ChildForm(instance=child)
    return render(request, 'crm/child_edit.html', {'form': form, 'child': child, 'page': 'child_edit'})


@login_required
def child_create_view(request):
    if request.method == 'POST':
        form = ChildForm(request.POST, request.FILES)
        if form.is_valid():
            child = form.save()
            messages.success(request, f'Ребенок {child.last_name} {child.first_name} добавлен')
            return redirect('child_card', child_id=child.id)
    else:
        form = ChildForm()
    return render(request, 'crm/child_edit.html', {'form': form, 'page': 'child_create'})

@login_required
def child_delete_view(request, child_id):
    child = get_object_or_404(Child, id=child_id)
    if request.method == 'POST':
        child_name = f"{child.last_name} {child.first_name}"
        child.delete()
        messages.success(request, f'Ребенок {child_name} удален')
        return redirect('attendance')
    return render(request, 'crm/child_delete.html', {'child': child, 'page': 'child_delete'})



@login_required
def add_subscription_view(request, child_id):
    child = get_object_or_404(Child, id=child_id)
    if request.method == 'POST':
        form = SubscriptionForm(request.POST)
        if form.is_valid():
            sub = form.save(commit=False)
            sub.child = child
            sub.save()
            messages.success(request, 'Абонемент добавлен')
            return redirect('child_card', child_id=child.id)
    else:
        form = SubscriptionForm()
    return render(request, 'crm/add_subscription.html', {'form': form, 'child': child, 'page': 'add_subscription'})


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Trainer, Group, ScheduleSlot
from .forms import TrainerForm, GroupForm, ScheduleSlotFormSet

# ==================== ТРЕНЕРЫ ====================

@login_required
def trainer_list_view(request):
    """Список тренеров"""
    trainers = Trainer.objects.all().prefetch_related('groups').order_by('full_name')
    context = {
        'trainers': trainers,
        'title': 'Тренеры',
        'subtitle': 'Управление тренерским составом',
        'page': 'trainers'
    }
    return render(request, 'crm/trainers.html', context)

@login_required
def trainer_create_view(request):
    """Создание тренера"""
    if request.method == 'POST':
        form = TrainerForm(request.POST)
        if form.is_valid():
            trainer = form.save()
            messages.success(request, f'Тренер {trainer.full_name} добавлен')
            return redirect('trainer_list')
    else:
        form = TrainerForm()

    context = {
        'form': form,
        'title': 'Новый тренер',
        'page': 'trainers'
    }
    return render(request, 'crm/trainer_edit.html', context)

@login_required
def trainer_edit_view(request, pk):
    """Редактирование тренера"""
    trainer = get_object_or_404(Trainer, pk=pk)
    if request.method == 'POST':
        form = TrainerForm(request.POST, instance=trainer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Данные тренера обновлены')
            return redirect('trainer_list')
    else:
        form = TrainerForm(instance=trainer)

    context = {
        'form': form,
        'trainer': trainer,
        'title': f'Редактирование: {trainer.full_name}',
        'page': 'trainers'
    }
    return render(request, 'crm/trainer_edit.html', context)

@login_required
def trainer_delete_view(request, pk):
    """Удаление тренера"""
    trainer = get_object_or_404(Trainer, pk=pk)
    if request.method == 'POST':
        # Проверяем, есть ли группы у тренера
        if trainer.groups.exists():
            messages.error(request, f'Нельзя удалить тренера {trainer.full_name}: у него есть группы. Сначала удалите или переназначьте группы.')
            return redirect('trainer_list')
        trainer.delete()
        messages.success(request, f'Тренер {trainer.full_name} удален')
        return redirect('trainer_list')

    context = {
        'trainer': trainer,
        'title': 'Удаление тренера',
        'page': 'trainers'
    }
    return render(request, 'crm/trainer_delete.html', context)


# ==================== ГРУППЫ ====================

@login_required
def group_list_view(request):
    """Список групп"""
    groups = Group.objects.select_related('trainer').prefetch_related('schedule', 'children').order_by('name')
    context = {
        'groups': groups,
        'title': 'Группы',
        'subtitle': 'Управление группами и расписанием',
        'page': 'groups'
    }
    return render(request, 'crm/groups.html', context)

@login_required
def group_create_view(request):
    """Создание группы с расписанием"""
    if request.method == 'POST':
        group_form = GroupForm(request.POST)
        slot_formset = ScheduleSlotFormSet(request.POST)

        if group_form.is_valid() and slot_formset.is_valid():
            group = group_form.save()
            slot_formset.instance = group
            slot_formset.save()
            messages.success(request, f'Группа "{group.name}" создана')
            return redirect('group_list')
    else:
        group_form = GroupForm()
        slot_formset = ScheduleSlotFormSet()

    context = {
        'group_form': group_form,
        'slot_formset': slot_formset,
        'title': 'Новая группа',
        'page': 'groups'
    }
    return render(request, 'crm/group_edit.html', context)

@login_required
def group_edit_view(request, pk):
    """Редактирование группы с расписанием"""
    group = get_object_or_404(Group, pk=pk)
    if request.method == 'POST':
        group_form = GroupForm(request.POST, instance=group)
        slot_formset = ScheduleSlotFormSet(request.POST, instance=group)

        if group_form.is_valid() and slot_formset.is_valid():
            group_form.save()
            slot_formset.save()
            messages.success(request, f'Группа "{group.name}" обновлена')
            return redirect('group_list')
    else:
        group_form = GroupForm(instance=group)
        slot_formset = ScheduleSlotFormSet(instance=group)

    context = {
        'group_form': group_form,
        'slot_formset': slot_formset,
        'group': group,
        'title': f'Редактирование: {group.name}',
        'page': 'groups'
    }
    return render(request, 'crm/group_edit.html', context)

@login_required
def group_delete_view(request, pk):
    """Удаление группы"""
    group = get_object_or_404(Group, pk=pk)
    if request.method == 'POST':
        children_count = group.children.count()
        if children_count > 0:
            messages.error(request, f'Нельзя удалить группу "{group.name}": в ней {children_count} детей. Сначала переведите детей в другие группы.')
            return redirect('group_list')
        group.delete()
        messages.success(request, f'Группа "{group.name}" удалена')
        return redirect('group_list')

    context = {
        'group': group,
        'title': 'Удаление группы',
        'page': 'groups'
    }
    return render(request, 'crm/group_delete.html', context)


from datetime import date, timedelta, datetime
from decimal import Decimal
from django.db.models import Sum, Count, Q
from django.http import HttpResponse

def _month_range(request):
    today = timezone.localdate()
    month_str = request.GET.get('month')
    try:
        month_start = datetime.strptime(month_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        month_start = today.replace(day=1)
    if month_start.month == 12:
        month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)
    return month_start, month_end, today

def _sessions_held(group, month_start, month_end, today):
    """Сколько занятий прошло по расписанию группы в месяце (до сегодня)."""
    weekdays = set(ScheduleSlot.objects.filter(group=group).values_list('weekday', flat=True))
    last = min(month_end, today)
    count, d = 0, month_start
    while d <= last:
        if d.weekday() in weekdays:
            count += 1
        d += timedelta(days=1)
    return count

def build_salary_data(month_start, month_end):
    """Собирает таблицу ЗП в формате как в Excel."""
    summary_rows, grand_total = [], 0
    trainers = []

    for trainer in Trainer.objects.prefetch_related('groups', 'salary_adjustments'):
        rows, total = [], 0
        for group in trainer.groups.all():
            rate = group.salary_rate or 0
            visits = Attendance.objects.filter(
                child__group=group, status='present',
                date__gte=month_start, date__lte=month_end).count()
            summa = Decimal(rate) * visits
            rows.append({'name': group.name, 'rate': rate, 'visits': visits, 'total': summa})
            summary_rows.append({'name': group.name, 'rate': rate, 'visits': visits, 'total': summa})
            total += summa

        for adj in trainer.salary_adjustments.filter(month=month_start):
            rows.append({'name': adj.title, 'rate': '', 'visits': '', 'total': adj.amount})
            summary_rows.append({'name': adj.title, 'rate': '', 'visits': '', 'total': adj.amount})
            total += adj.amount

        if rows:
            trainers.append({'trainer': trainer, 'rows': rows, 'total': total})
            grand_total += total

    return {'summary_rows': summary_rows, 'grand_total': grand_total, 'trainers': trainers}


@login_required
def statistics_view(request):
    month_start, month_end, today = _month_range(request)

    children = Child.objects.all()
    total_children = children.count()
    active_children = children.filter(status=Child.Status.ACTIVE).count()

    new_qs = children.filter(created_at__date__gte=month_start, created_at__date__lte=month_end)
    new_count = new_qs.count()
    new_kept = new_qs.filter(status=Child.Status.ACTIVE).count()
    left_count = children.filter(
        status__in=[Child.Status.LOST, Child.Status.ARCHIVED],
        archived_at__gte=month_start, archived_at__lte=month_end).count()

    # Выручка
    revenue_today = Payment.objects.filter(date=today).aggregate(s=Sum('amount'))['s'] or 0
    revenue_month = Payment.objects.filter(
        date__gte=month_start, date__lte=month_end).aggregate(s=Sum('amount'))['s'] or 0
    # Прогноз: факт месяца + ожидаемые продления (абонементы, кончающиеся до конца месяца)
    expected = Subscription.objects.filter(
        end_date__gte=today, end_date__lte=month_end).aggregate(s=Sum('price'))['s'] or 0
    potential = Decimal(revenue_month) + Decimal(expected)

    target = RevenueTarget.objects.filter(month=month_start).first()

    # По группам: спортсмены, посещения, посещаемость
    groups_stats = []
    for g in Group.objects.filter(is_active=True).select_related('trainer'):
        kids = g.children.count()
        present = Attendance.objects.filter(
            child__group=g, status='present',
            date__gte=month_start, date__lte=month_end).count()
        absent = Attendance.objects.filter(
            child__group=g, status='absent',
            date__gte=month_start, date__lte=month_end).count()
        sessions = _sessions_held(g, month_start, month_end, today)
        capacity = kids * sessions
        attendance_pct = round(present * 100 / capacity) if capacity else 0
        groups_stats.append({
            'group': g, 'kids': kids, 'present': present, 'absent': absent,
            'sessions': sessions, 'attendance_pct': attendance_pct,
        })

    # По тренерам: ушедшие + посещаемость
    trainers_stats = []
    for t in Trainer.objects.filter(is_active=True):
        t_left = children.filter(
            group__trainer=t, status__in=[Child.Status.LOST, Child.Status.ARCHIVED],
            archived_at__gte=month_start, archived_at__lte=month_end).count()
        t_groups = [gs for gs in groups_stats if gs['group'].trainer_id == t.id]
        t_present = sum(gs['present'] for gs in t_groups)
        t_capacity = sum(gs['kids'] * gs['sessions'] for gs in t_groups)
        trainers_stats.append({
            'trainer': t,
            'left': t_left,
            'present': t_present,
            'attendance_pct': round(t_present * 100 / t_capacity) if t_capacity else 0,
        })

    context = {
        'month_start': month_start, 'month_end': month_end, 'today': today,
        'total_children': total_children, 'active_children': active_children,
        'new_count': new_count, 'new_kept': new_count - 0 and new_kept,
        'left_count': left_count,
        'revenue_today': revenue_today, 'revenue_month': revenue_month,
        'potential': potential, 'target': target,
        'groups_stats': groups_stats, 'trainers_stats': trainers_stats,
        'title': 'Статистика', 'page': 'statistics',
    }
    return render(request, 'crm/statistics.html', context)


@login_required
def salaries_view(request):
    month_start, month_end, today = _month_range(request)
    data = build_salary_data(month_start, month_end)
    context = {**data, 'month_start': month_start, 'month_end': month_end,
               'title': 'ЗП тренеров', 'page': 'salaries'}
    return render(request, 'crm/salaries.html', context)


@login_required
def salaries_export_view(request):
    """Выгрузка таблицы ЗП в Excel (формат как в ручном подсчете)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    month_start, month_end, today = _month_range(request)
    data = build_salary_data(month_start, month_end)

    wb = Workbook()
    ws = wb.active
    ws.title = 'ЗП тренеров'

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill('solid', fgColor='92D050')
    gray = PatternFill('solid', fgColor='E2EFDA')
    bold = Font(bold=True)
    center = Alignment(horizontal='center')

    ws.merge_cells('A1:D1')
    ws['A1'] = f"{month_start:%d.%m.%Y} — {month_end:%d.%m.%Y}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    row = 3
    for header, col in [('Группа', 'A'), ('Ставка', 'B'), ('Посещение', 'C'), ('Всего', 'D')]:
        c = ws[f'{col}{row}']; c.value = header; c.font = bold; c.border = border; c.fill = gray
    row += 1

    for r in data['summary_rows']:
        ws[f'A{row}'] = r['name']; ws[f'B{row}'] = r['rate']
        ws[f'C{row}'] = r['visits']; ws[f'D{row}'] = r['total']
        for col in 'ABCD': ws[f'{col}{row}'].border = border
        row += 1

    ws[f'A{row}'] = 'Итого'; ws[f'A{row}'].font = bold
    ws[f'D{row}'] = data['grand_total']; ws[f'D{row}'].font = bold
    ws[f'D{row}'].fill = gray
    for col in 'ABCD': ws[f'{col}{row}'].border = border
    row += 2

    for t in data['trainers']:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = t['trainer'].full_name
        ws[f'A{row}'].font = bold; ws[f'A{row}'].fill = green
        for col in 'BCD': ws[f'{col}{row}'].fill = green
        row += 1
        for r in t['rows']:
            ws[f'A{row}'] = r['name']; ws[f'B{row}'] = r['rate']
            ws[f'C{row}'] = r['visits']; ws[f'D{row}'] = r['total']
            for col in 'ABCD': ws[f'{col}{row}'].border = border
            row += 1
        ws[f'C{row}'] = 'Итого'; ws[f'C{row}'].font = bold
        ws[f'D{row}'] = t['total']; ws[f'D{row}'].font = bold
        row += 2

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="salaries_{month_start:%Y-%m}.xlsx"'
    wb.save(response)
    return response


@login_required
def payments_table_view(request):
    month_start, month_end, today = _month_range(request)
    payments = Payment.objects.filter(
        date__gte=month_start, date__lte=month_end
    ).select_related('child', 'child__group', 'created_by', 'subscription').order_by('-date')
    total = payments.aggregate(s=Sum('amount'))['s'] or 0
    context = {'payments': payments, 'total': total,
               'month_start': month_start, 'month_end': month_end,
               'title': 'Предварительные оплаты', 'page': 'payments_table'}
    return render(request, 'crm/payments_table.html', context)


@login_required
def expenses_table_view(request):
    month_start, month_end, today = _month_range(request)
    if request.method == 'POST':
        title = request.POST.get('title')
        amount = request.POST.get('amount')
        if title and amount:
            Expense.objects.create(title=title, amount=amount, created_by=request.user)
            messages.success(request, 'Расход добавлен')
            return redirect('expenses_table')
    expenses = Expense.objects.filter(
        date__gte=month_start, date__lte=month_end).order_by('-date')
    total = expenses.aggregate(s=Sum('amount'))['s'] or 0
    context = {'expenses': expenses, 'total': total,
               'month_start': month_start, 'month_end': month_end,
               'title': 'Расходы', 'page': 'expenses_table'}
    return render(request, 'crm/expenses_table.html', context)
